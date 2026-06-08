"""Per-tool exports: CGI workbook (Substance + Env Refs) and TwelveLabs rows with
recommendation role."""

import io

from openpyxl import load_workbook

from director_agent.draftstore.store import DraftRecord
from director_agent.outputs.exports import build_cgi_workbook, build_twelvelabs_rows, twelvelabs_csv


def _rec(cell_id, cell_type, outputs, approved=True):
    return DraftRecord(cell_id=cell_id, cell_type=cell_type, review_status="auto_accept",
                       approved=approved, envelope={"outputs": outputs}, created_at="t")


SUBSTANCE = {"type": "substance_row", "confidence": "high", "row": {
    "nameplate": "Ram 1500", "specific_trim_request": "D28H91 - Tradesman (Base)",
    "location_variant": "Great Lakes", "color_preference_hex": "#B61A22",
    "camera_angles": "action_front_3q", "ai_image_generator_prompt": "360 HDRI, no people."}}
CG_ENV = {"type": "cg_env_prompt", "confidence": "medium", "for_pipeline": "runway", "prompt": "winter env"}
TL = {"type": "twelvelabs_query", "confidence": "high",
      "query": {"tags": ["ram_1500", "badge"], "natural_language": "Ram 1500 hemi badge close-up, winter daylight"}}
STOCK = {"type": "stock_search", "confidence": "medium",
         "natural_language_description": "snowy town at dusk", "tags_for_indexed_search": ["winter", "no_people"]}


def test_cgi_workbook_has_both_sheets_with_recommendation():
    recs = [_rec("s5__great_lakes", "regionalized_ai_scenes", [SUBSTANCE]),
            _rec("s1__great_lakes", "regionalized_running_w_cgi_ai", [CG_ENV])]
    wb = load_workbook(io.BytesIO(build_cgi_workbook(recs)))
    assert wb.sheetnames == ["Substance Variants", "Env Refs"]
    sub = wb["Substance Variants"]
    assert [c.value for c in sub[1]] == [  # strict 6-col format, unchanged
        "Nameplate", "Specific Trim Request", "Location Variant",
        "Color Preference (HEX)", "Camera Angles", "AI Image Generator Prompt"]
    assert sub.cell(row=2, column=4).value == "#B61A22"
    env = wb["Env Refs"]
    assert env.cell(row=1, column=3).value == "Recommendation"
    assert "Regionalize" in env.cell(row=2, column=3).value  # headline present
    assert env.cell(row=2, column=5).value == "winter env"  # prompt now col 5


def test_twelvelabs_rows_carry_role():
    rows = build_twelvelabs_rows([
        _rec("s3__great_lakes", "existing_running_footage", [TL]),
        _rec("s1__great_lakes", "regionalized_running_w_cgi_ai", [TL]),  # base plate -> supporting
        _rec("s2__southwest", "stock", [STOCK]),
    ])
    by_cell = {r["cell_id"]: r for r in rows}
    assert by_cell["s3__great_lakes"]["role"] == "primary"
    assert "Core footage" in by_cell["s3__great_lakes"]["recommendation"]
    assert by_cell["s1__great_lakes"]["role"] == "supporting"   # regionalized base plate
    assert by_cell["s2__southwest"]["role"] == "primary"
    assert {r["query_type"] for r in rows} == {"footage", "stock"}


def test_csv_header_has_role():
    csv_text = twelvelabs_csv([_rec("a__great_lakes", "existing_running_footage", [TL])])
    assert "role,recommendation,query_type" in csv_text
