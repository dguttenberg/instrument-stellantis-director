"""Full South:15 run across GL + SW with a fake director (no API). Verifies the
runner orchestration: every scene emits, continuity threads, the regionalized_ai
cell writes a Substance row per lane, and GL vs SW resolve different environment
content."""

from openpyxl import load_workbook

from director_agent.bg import build_client
from director_agent.config import DATA_DIR
from director_agent.director.cell_specs import buckets_for
from director_agent.draftstore import LocalDraftStore
from director_agent.pipeline import PipelineRunner, load_dials, load_script, load_styling
from director_agent.schemas.bg import BGRequest
from director_agent.schemas.cell import CellOutputEnvelope


class FakeDirector:
    """Returns a canned, valid envelope per cell type; varies hex by lane so GL≠SW.
    Records the prior_cell_resolved it received to prove continuity threading."""

    def __init__(self):
        self.seen_prior = []

    def run(self, cell, slices, dials, styling, prior_cell_resolved=None):
        self.seen_prior.append(prior_cell_resolved)
        hex_by_lane = {"great_lakes": "#B61A22", "southwest": "#FFFFFF"}
        outputs = []
        if cell.cell_type == "regionalized_ai_scenes":
            outputs.append(
                {
                    "type": "substance_row",
                    "confidence": "high",
                    "row": {
                        "Nameplate": "Ram 1500",
                        "Specific Trim Request": "D28H91 - Tradesman (Base)",
                        "Location Variant": "Great Lakes" if cell.lane == "great_lakes" else "Southwest",
                        "Color Preference (HEX)": hex_by_lane[cell.lane],
                        "Camera Angles": "action_front_3q, hero_front",
                        "AI Image Generator Prompt": "A photorealistic 360 HDRI ... no people.",
                    },
                }
            )
        elif cell.cell_type == "stock":
            outputs.append({"type": "stock_search", "confidence": "medium",
                            "natural_language_description": "regional establishing shot",
                            "tags_for_indexed_search": ["winter", "no_people"]})
        else:  # running footage variants
            outputs.append({"type": "twelvelabs_query", "confidence": "high",
                            "query": {"tags": ["ram_1500"], "natural_language": "ram driving"}})
        if cell.super_called:
            outputs.append({"type": "super_text", "confidence": "high",
                            "content": cell.super_intent or "", "voice_tags": ["edgy_bold_direct"]})
        return CellOutputEnvelope(cell_id=cell.cell_id, cell_type=cell.cell_type, outputs=outputs)


def test_full_south15_both_lanes(tmp_path):
    script = load_script(DATA_DIR / "scripts" / "map_retail_ram_test_15_south.json")
    dials = load_dials(DATA_DIR / "dials" / "year_end_demo.json")
    styling = load_styling(DATA_DIR / "styling" / "year_end_demo.json")
    store = LocalDraftStore(tmp_path / "drafts.sqlite")
    substance = tmp_path / "substance.xlsx"

    for lane in ["great_lakes", "southwest"]:
        director = FakeDirector()
        runner = PipelineRunner(build_client(), director, store, substance, season="winter")
        result = runner.run(script, lane=lane, dials=dials, styling=styling)

        assert len(result.envelopes) == 6
        # Continuity: scene 1 gets no prior; every later scene gets the previous envelope.
        assert director.seen_prior[0] is None
        assert director.seen_prior[1]["cell_id"].startswith("south15_scene1")
        assert director.seen_prior[5]["cell_id"].startswith("south15_scene5")

    # One Substance row per lane (scene 5 only) -> header + 2 rows.
    ws = load_workbook(str(substance)).active
    assert ws.max_row == 3
    hexes = {ws.cell(row=r, column=4).value for r in (2, 3)}
    assert hexes == {"#B61A22", "#FFFFFF"}  # GL red vs SW white -> visibly different

    # 12 drafts total (6 scenes x 2 lanes).
    assert len(store.list()) == 12


def test_gl_and_sw_resolve_different_environment():
    client = build_client()

    def env(lane):
        req = BGRequest(
            request_id="x", brand="ram", lane=lane,
            buckets_needed=buckets_for("regionalized_ai_scenes", "D28H91", "winter"),
        )
        return client.resolve(req).slices[f"environmental_context.{lane}"].content

    gl = env("great_lakes")
    sw = env("southwest")
    assert gl["weather"] != sw["weather"]
    assert gl["palette_hints"] != sw["palette_hints"]
