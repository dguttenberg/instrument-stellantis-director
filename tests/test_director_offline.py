"""Vertical slice without a live API: BG resolve -> Director (fake client) ->
envelope -> Substance xlsx -> draft store. Proves all plumbing but the round-trip."""

from openpyxl import load_workbook

from director_agent.director.director import Director
from director_agent.draftstore import LocalDraftStore
from director_agent.outputs.substance_excel import append_substance_rows
from director_agent.schemas.cell import SubstanceRow

from conftest import FakeAnthropic, make_cell, resolve_slices

# Canned director emission (what the model would return for scene 5, GL).
EMISSION = {
    "outputs": [
        {
            "type": "substance_row",
            "confidence": "high",
            "row": {
                "Nameplate": "Ram 1500",
                "Specific Trim Request": "D28H91 - Tradesman (Base)",
                "Location Variant": "Great Lakes",
                "Color Preference (HEX)": "#B61A22",
                "Camera Angles": "action_front_3q, hero_front",
                "AI Image Generator Prompt": (
                    "A photorealistic 360 HDRI of a Great Lakes winter rural two-lane road, "
                    "sustained snow, lake-effect overcast gray sky, bare deciduous, salt-streaked "
                    "asphalt, flat low-angle late-afternoon light; pole barn in mid-distance; no people."
                ),
            },
            "director_notes": "Red reads warm against the cool Great Lakes winter palette.",
        },
        {
            "type": "super_text",
            "confidence": "high",
            "content": "Get the high output 6.7L Cummins turbo diesel.",
            "voice_tags": ["edgy_bold_direct"],
        },
    ],
    "gaps_flagged": [],
}


def test_vertical_slice_offline(tmp_path, demo_dials, demo_styling):
    cell = make_cell("regionalized_ai_scenes")
    slices = resolve_slices(cell)
    assert not slices.gaps_flagged  # all 7 buckets resolve for GL

    director = Director(client=FakeAnthropic(EMISSION))
    envelope = director.run(cell, slices, demo_dials, demo_styling)

    assert envelope.cell_id == cell.cell_id
    assert envelope.cell_type == "regionalized_ai_scenes"
    assert envelope.draft is True
    types = [o.type for o in envelope.outputs]
    assert "substance_row" in types and "super_text" in types

    # Substance row drops into the demo xlsx format.
    rows = [o for o in envelope.outputs if isinstance(o, SubstanceRow)]
    path = append_substance_rows(rows, tmp_path / "out.xlsx")
    ws = load_workbook(str(path)).active
    assert ws.cell(row=2, column=3).value == "Great Lakes"
    assert ws.cell(row=2, column=4).value == "#B61A22"

    # Draft store: all-high -> auto_accept + approved by default (spec §7).
    store = LocalDraftStore(tmp_path / "drafts.sqlite")
    rec = store.put(envelope)
    assert rec.review_status == "auto_accept"
    assert rec.approved is True
    assert store.get(cell.cell_id).cell_type == "regionalized_ai_scenes"


def test_corrective_retry_sends_tool_result_turn(demo_dials, demo_styling):
    """A first-attempt validation miss must retry with a tool_result block after the
    assistant tool_use turn (API requirement), not a bare user text message."""
    cell = make_cell("regionalized_ai_scenes")
    slices = resolve_slices(cell)
    invalid = {"outputs": [{"type": "super_text", "confidence": "loud"}]}  # bad confidence enum
    fake = FakeAnthropic([invalid, EMISSION])

    director = Director(client=fake)
    env = director.run(cell, slices, demo_dials, demo_styling)

    assert fake.calls == 2  # one miss, one corrective retry
    # The retry's message history must end: assistant(tool_use) -> user(tool_result).
    retry_messages = fake.received_messages[1]
    assert retry_messages[-2]["role"] == "assistant"
    last = retry_messages[-1]
    assert last["role"] == "user"
    assert last["content"][0]["type"] == "tool_result"
    assert last["content"][0]["is_error"] is True
    # And it ultimately produced the valid envelope.
    assert any(o.type == "substance_row" for o in env.outputs)


def test_stringified_outputs_are_coerced(demo_dials, demo_styling):
    """Opus tool-use sometimes serializes `outputs` as a JSON string; the director
    must still produce a valid envelope (regression: rs=0.2 live failure)."""
    import json as _json

    cell = make_cell("regionalized_ai_scenes")
    slices = resolve_slices(cell)
    stringified = {"outputs": _json.dumps(EMISSION["outputs"]), "gaps_flagged": "[]"}
    director = Director(client=FakeAnthropic(stringified))
    env = director.run(cell, slices, demo_dials, demo_styling)
    assert any(o.type == "substance_row" for o in env.outputs)


def test_dial_directives_change_system_prompt(demo_styling):
    """The regional_specificity dial must change the director's instructions —
    the pitch mechanic. Low vs high produce visibly different directives."""
    from director_agent.director.prompt import build_system_prompt
    from director_agent.schemas.dials import Dials

    low = build_system_prompt(demo_styling, Dials(regional_specificity=0.2), "regionalized_ai_scenes")
    high = build_system_prompt(demo_styling, Dials(regional_specificity=0.9), "regionalized_ai_scenes")
    assert "Regional specificity LOW" in low
    assert "Regional specificity HIGH" in high
    assert low != high


def test_disallowed_output_type_is_dropped_and_flagged(demo_dials, demo_styling):
    # existing_running_footage may not emit substance_row -> should be dropped + flagged.
    cell = make_cell("existing_running_footage")
    slices = resolve_slices(cell)
    bad = {
        "outputs": [
            {"type": "substance_row", "confidence": "high", "row": EMISSION["outputs"][0]["row"]},
            {"type": "twelvelabs_query", "confidence": "high",
             "query": {"tags": ["ram_1500"], "natural_language": "tight close-up of HEMI badge"}},
        ],
        "gaps_flagged": [],
    }
    director = Director(client=FakeAnthropic(bad))
    env = director.run(cell, slices, demo_dials, demo_styling)
    types = [o.type for o in env.outputs]
    assert "substance_row" not in types
    assert "twelvelabs_query" in types
    assert any("disallowed" in g for g in env.gaps_flagged)
