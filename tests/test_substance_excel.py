"""Substance row drops into the exact 6-column xlsx format."""

from openpyxl import load_workbook

from director_agent.outputs.substance_excel import (
    SUBSTANCE_COLUMNS,
    append_substance_rows,
    write_substance_workbook,
)
from director_agent.schemas.cell import SubstanceRow


def _row(location="Great Lakes", hex_="#B61A22"):
    return SubstanceRow.model_validate(
        {
            "type": "substance_row",
            "confidence": "high",
            "row": {
                "Nameplate": "Ram 1500",
                "Specific Trim Request": "D28H91 - Tradesman (Base)",
                "Location Variant": location,
                "Color Preference (HEX)": hex_,
                "Camera Angles": "action_front_3q, hero_front",
                "AI Image Generator Prompt": "A photorealistic 360 HDRI ... no people.",
            },
        }
    )


def test_write_then_read_matches_demo_format(tmp_path):
    path = tmp_path / "substance.xlsx"
    write_substance_workbook([_row()], path)
    wb = load_workbook(str(path))
    ws = wb.active
    header = [c.value for c in ws[1]]
    assert header == SUBSTANCE_COLUMNS
    first = [c.value for c in ws[2]]
    assert first[0] == "Ram 1500"
    assert first[3] == "#B61A22"


def test_append_creates_then_extends(tmp_path):
    path = tmp_path / "substance.xlsx"
    append_substance_rows([_row(location="Great Lakes")], path)
    append_substance_rows([_row(location="Southwest", hex_="#FFFFFF")], path)
    wb = load_workbook(str(path))
    ws = wb.active
    assert ws.max_row == 3  # header + 2 rows
    assert ws.cell(row=3, column=3).value == "Southwest"
