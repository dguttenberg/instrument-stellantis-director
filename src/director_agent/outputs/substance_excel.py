"""Write substance_row outputs into the exact column shape of the demo spreadsheet
(Substance Automation-ram_variant_demo.xlsx). The row drops in with no translation."""

from __future__ import annotations

from pathlib import Path
from typing import Iterable, List

from openpyxl import Workbook, load_workbook

from ..schemas.cell import SubstanceRow

# Exact header order, verified against _ref/Substance Automation-ram_variant_demo.xlsx.
SUBSTANCE_COLUMNS = [
    "Nameplate",
    "Specific Trim Request",
    "Location Variant",
    "Color Preference (HEX)",
    "Camera Angles",
    "AI Image Generator Prompt",
]


def _row_values(row: SubstanceRow) -> List[str]:
    d = row.row.model_dump(by_alias=True)
    return [d[col] for col in SUBSTANCE_COLUMNS]


def write_substance_workbook(rows: Iterable[SubstanceRow], path: Path) -> Path:
    """Create (overwrite) a workbook with the header + the given rows."""
    path = Path(path)
    path.parent.mkdir(parents=True, exist_ok=True)
    wb = Workbook()
    ws = wb.active
    ws.title = "Variants"
    ws.append(SUBSTANCE_COLUMNS)
    for row in rows:
        ws.append(_row_values(row))
    wb.save(str(path))
    return path


def append_substance_rows(rows: Iterable[SubstanceRow], path: Path) -> Path:
    """Append rows to an existing workbook, creating it (with header) if absent."""
    path = Path(path)
    if not path.exists():
        return write_substance_workbook(rows, path)
    wb = load_workbook(str(path))
    ws = wb.active
    for row in rows:
        ws.append(_row_values(row))
    wb.save(str(path))
    return path
